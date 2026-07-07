# Genera src/catalogo.ts desde el Excel "Buscador Repuestos SAP por Modo de Falla.xlsx".
# Re-correr cuando el Excel cambie:  python scripts/gen_catalogo.py
import sys, io, json, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

XLSX = r"C:\Users\braya\OneDrive\Escritorio\Buscador Repuestos SAP por Modo de Falla.xlsx"
OUT = os.path.join(os.path.dirname(__file__), "..", "src", "catalogo.ts")

wb = openpyxl.load_workbook(XLSX, data_only=True)

def s(v):
    return "" if v is None else str(v).strip()

# --- MAESTRO: MF -> repuestos ---
maestro = []
for row in list(wb["MAESTRO"].iter_rows(values_only=True))[1:]:
    mf = s(row[0])
    if not mf:
        continue
    cant_raw = s(row[5])
    try:
        cant = int(float(cant_raw)) if cant_raw else 1
    except ValueError:
        cant = 1
    maestro.append({
        "mf": mf, "falla": s(row[1]), "planta": s(row[2]),
        "componente": s(row[3]), "sap": s(row[4]),
        "cantidad": cant, "conjunto": s(row[6]),
    })

# --- CATALOGO: 109 materiales ---
catalogo = []
for row in list(wb["CATALOGO"].iter_rows(values_only=True))[1:]:
    sap = s(row[0])
    if not sap:
        continue
    catalogo.append({"sap": sap, "nombre": s(row[1]), "planta": s(row[2]), "conjunto": s(row[3])})

# --- 14 modos de falla ---
nombres = {
    "MF1": "Tapón", "MF2": "Canastillo", "MF3": "Piedmont", "MF4": "Sideport",
    "MF5": "Americana", "MF6": "Brazo", "MF7": "Tubing", "MF8": "Manifold",
    "MF9": "Tapa", "MF10": "Flange", "MF11": "Piting", "MF12": "Venteo Alta",
    "MF13": "Venteo Baja", "MF14": "Tapón manifold",
}
modos = [{"codigo": k, "nombre": v} for k, v in nombres.items()]

def js(arr):
    return json.dumps(arr, ensure_ascii=False, indent=2)

ts = '''// AUTO-GENERADO desde "Buscador Repuestos SAP por Modo de Falla.xlsx"
// No editar a mano: re-generar con  python scripts/gen_catalogo.py

export interface MaestroFila {
  mf: string
  falla: string
  planta: string
  componente: string
  sap: string
  cantidad: number
  conjunto: string
}

export interface CatalogoItem {
  sap: string
  nombre: string
  planta: string
  conjunto: string
}

export interface ModoFalla { codigo: string; nombre: string }

export const PLANTAS_RO: string[] = ["EWS", "EWSE", "Planta 0"]

export const MODOS_FALLA: ModoFalla[] = ''' + js(modos) + '''

export const MAESTRO: MaestroFila[] = ''' + js(maestro) + '''

export const CATALOGO: CatalogoItem[] = ''' + js(catalogo) + '''
'''

with open(OUT, "w", encoding="utf-8") as f:
    f.write(ts)

print(f"OK -> {os.path.normpath(OUT)}")
print(f"maestro={len(maestro)}  catalogo={len(catalogo)}  modos={len(modos)}")
